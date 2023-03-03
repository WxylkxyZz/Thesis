from User_Agents import user_agents
import proxy_operation
import asyncio
import aiohttp
import random
import re
import json
import openpyxl

headers = {
    ":authority": "club.jd.com",
    ":scheme": "https",
    "Accept": "*/*",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Connection": "keep-alive",
    # "Cookie": "token=9a5fb93299b903308394b4289b3c86d7,1,931225; __tk=SkrIVUJATLYDWcbIiDVIikrJTkS2iUlDSDfLWcaCTIaDVAYCiIY2VY,1,931225; shshshfp=ea533145daabc9b06f1b33a9cf43ad81; shshshfpa=8920ebb7-6af2-bbaf-7ef6-2265949ad405-1676205048; shshshfpx=8920ebb7-6af2-bbaf-7ef6-2265949ad405-1676205048; shshshsID=4528b18f4fbff628b14243dc03dfd676_1_1676205048434; jsavif=1; __jda=122270672.16762050486901085966906.1676205049.1676205049.1676205049.1; __jdb=122270672.1.16762050486901085966906|1.1676205049; __jdc=122270672; __jdv=122270672|direct|-|none|-|1676205048690; shshshfpb=kI_8t2BIx3CaIcWu0LZrYwQ; areaId=5; __jdu=16762050486901085966906; ip_cityCode=142; ipLoc-djd=5-274-49708-49891; 3AB9D23F7A4B3C9B=MPFYZXBXVF4EW5W4JT6UY3JJ2GAIM2GGDBG65A36BIASHZTK3MKRQ7HILIBXALAC5BUB3ZLFSAYOHHQYHOS3J2DC34",
    "Cookie": "__jdu=1677768678870785055026; areaId=5; PCSYCityID=CN_130000_131100_0; shshshfpa=daa56b13-f25b-5541-a186-b71c7f86f4e3-1677768679; shshshfpx=daa56b13-f25b-5541-a186-b71c7f86f4e3-1677768679; shshshfpb=g8B2i-vIOScLB1oPRAKUsyg; ipLoc-djd=5-275-296-41561; jwotest_product=99; unpl=JF8EAMdnNSttWB5UABMKExMSG1gGW14BSURXbGAEBlRQSVJWE1ZMQRd7XlVdXxRKFR9uZRRVVFNOVw4aAysSEXteXVdZDEsWC2tXVgQFDQ8VXURJQlZAFDNVCV9dSRZRZjJWBFtdT1xWSAYYRRMfDlAKDlhCR1FpMjVkXlh7VAQrAhwUFk1aXVlbAUIRAmluAVVZXEpXAisDKxUge21TWV0JQh4zblcEZB8MF1YMGQccXxBMW1JYWgFMEQpmYQRSVFxKUAEaARwiEXte; __jdv=76161171|baidu-pinzhuan|t_288551095_baidupinzhuan|cpc|0f3d30c8dba7459bb52f2eb5eba8ac7d_0_1d1499903a4c4283be260c8937b9efb6|1677804121456; jsavif=1; shshshfp=39d05cc62e4144f4819fbce8291319de; __jda=122270672.1677768678870785055026.1677768679.1677768679.1677804121.2; token=90a5209c716a3b152f99becbd270eaf1,2,932115; __tk=lvIZSYI6kEs0XYn1l1tphsbXlEPPYYIEgXk0nabXlsXJXuIElEXZhsbYkvG0j0bxgZfvSuMd,2,932115; __jdc=122270672; 3AB9D23F7A4B3C9B=RCLMVFZWZ4QTEZVOZLEOZDTEPNY6TSDVOWPCGVJYA34LM2POH25GTJZV22YTCV4OVOWUTTG74BINU7P4GVPO7FUQ2M; JSESSIONID=B6C0857B64CF70EC657B52542075D8CC.s1; shshshsID=9d8649a64fde5c8088a2de84ddf9d6b3_32_1677808387263; __jdb=122270672.33.1677768678870785055026|2.1677804121",
    "Host": "club.jd.com",
    "Referer": "https://item.jd.com/",
    "sec-ch-ua": '"Not_A Brand";v="99", "Google Chrome";v="109", "Chromium";v="109',
    'sec-ch-ua-mobile': '?0',
    'Sec-Fetch-Dest': 'script',
    "sec-ch-ua-platform": '"Windows"',
    'Sec-Fetch-Mode': 'no-cors',
    'Sec-Fetch-Site': 'same-site',
    "User-Agent": random.choice(user_agents)
}


async def get_params(ID, num, score):
    page_list = []
    for i in range(0, num):
        params = {
            'callback': 'fetchJSON_comment98',  # 固定值
            'productId': ID,  # 产品 ID
            'score': score,  # 0-3 分别对应全部评价,差评，中评，好评等
            'sortType': '5',  # 排序方式 5:默认排序 6:时间排序
            'page': i,  # 当前的评论页 从0开始
            'pageSize': '10',  # 每页评论的数量
            'isShadowSku': '0',
            'rid': '0',
            'fold': '1'
        }
        page_list.append(params)
    return page_list


# 代理池
proxies = []


async def Replenish():
    plist = proxy_operation.get_proxy()
    for proxy in plist:
        proxies.append(proxy)


async def judge_proxy_num():
    num = len(proxies)
    if num <= 5:
        await Replenish()


async def fetch(session, params, queue):
    async with semaphore:
        # base_url = "https://club.jd.com/comment/productPageComments.action?" 没有勾选只看当前商品评价 其他参数一样
        base_url = "https://club.jd.com/comment/skuProductPageComments.action?"
        proxy = random.choice(proxies)
        print(f"当前proxy: {proxy},发起请求的page: {params['page']}")
        async with session.get(base_url, headers=headers, params=params, proxy=proxy) as resp:
            if resp.status == 200:
                html = await resp.text()
                print(f"当前proxy: {proxy},已返回数据的page: {params['page']}")
                if html:
                    print(f"第{params['page']}页的html进入queue")
                    await queue.put(html)
                else:
                    print(f'没有返回内容 删除{proxy} 重新选择代理')
                    proxies.remove(proxy)
                    await judge_proxy_num()
                    await fetch(session, params, queue)
            else:
                print(f'代理池中代理无效 删除{proxy} 重新选择代理')
                proxies.remove(proxy)
                await judge_proxy_num()
                await fetch(session, params, queue)


async def producer(queue, page_list):
    async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=False)) as session:
        tasks = [asyncio.create_task(fetch(session, params, queue)) for params in page_list]
        await asyncio.wait(tasks)


async def create_sheet():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1).value = 'Amount'
    sheet.cell(row=1, column=2).value = 'UserName'
    sheet.cell(row=1, column=3).value = 'ID'
    sheet.cell(row=1, column=4).value = 'ReferenceName'
    sheet.cell(row=1, column=5).value = 'Score'
    sheet.cell(row=1, column=6).value = 'Comment'
    workbook.save('comments.xlsx')


async def consumer(queue):
    await create_sheet()
    wb = openpyxl.load_workbook('comments.xlsx')
    sheet = wb.active
    i = 1
    while True:
        html = await queue.get()
        if html is None:
            break
        else:
            html_json = re.search('(?<=fetchJSON_comment98\().*(?=\);)', html).group(0)
            comments_data = json.loads(html_json)["comments"]
            for comments in comments_data:
                sheet.cell(row=i + 1, column=1).value = i
                sheet.cell(row=i + 1, column=2).value = comments['nickname']
                sheet.cell(row=i + 1, column=3).value = comments['id']
                sheet.cell(row=i + 1, column=4).value = comments['referenceName']
                sheet.cell(row=i + 1, column=5).value = comments['score']
                sheet.cell(row=i + 1, column=6).value = comments['content']
                i += 1
    wb.save('comments.xlsx')


# 定义主函数
async def main():
    ID = 100019151854  # 商品ID
    num = 10  # 页数
    score = 1  # 评价等级
    page_list = await get_params(ID, num, score)  # 获取params, 每个是一页
    queue = asyncio.Queue()  # 创建队列
    await Replenish()  # 获取代理列表
    print(f"params以及代理ip获取完毕 进入生产者消费者模式！\n代理列表（{len(proxies)}）个", proxies)
    consumer_task = asyncio.create_task(consumer(queue))
    await producer(queue, page_list)
    await queue.put(None)  # 最后留None让Consumer停止
    await consumer_task
    print("生产者消费者都已完成！")


if __name__ == '__main__':
    semaphore = asyncio.Semaphore(10)
    asyncio.run(main())
    # 注 这么多几乎并发的要求从一个Cooike发出可能会被反爬即便使用了代理 可以调整30页30页爬
    # score=0 num = [0,99] 整100页, 10条/页
    # 以下均为500ml*15瓶
    # 青柑普洱茶  ID=100019151854 score=1 -> max(num) = 5   score=2 -> max(num) = 7
    # 茉莉花茶   score=1 -> max(num) = 20   score=2 -> max(num) = 22
    # 乌龙茶     score=1 -> max(num) = 11   score=2 -> max(num) = 5
    # 绿茶      score=1 -> max(num) = 8   score=2 -> max(num) = 16
    # 红茶      score=1 -> max(num) = 5   score=2 -> max(num) = 7
    # 中差评 106页 共1060条
    # 好评 500页 共5000条
